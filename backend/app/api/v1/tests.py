import io
import random
from datetime import datetime, timezone
from uuid import UUID

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.api.deps import require_roles
from app.db.session import get_db
from app.word_tests_import import parse_docx_to_questions
from app.models.quiz import (
    QuestionType,
    QuizAttempt,
    QuizAttemptAnswer,
    QuizOption,
    QuizQuestion,
    QuizTest,
    QuizTestAssignment,
)
from app.models.user import JobTitleCatalog, RestaurantCatalog, Role, User
from app.schemas.quiz import (
    QuizAnalyticsResponse,
    QuizAnalyticsSummary,
    QuizAttemptDetailPublic,
    QuizOptionTakePublic,
    QuizOptionPublic,
    QuizQuestionAnalyticsItem,
    QuizAttemptQuestionDetailPublic,
    QuizQuestionResultPublic,
    QuizQuestionTakePublic,
    QuizQuestionPublic,
    QuizScoreboardCell,
    QuizScoreboardResponse,
    QuizScoreboardTestRef,
    QuizScoreboardUser,
    QuizSubmitRequest,
    QuizSubmitResultPublic,
    QuizTestAssignmentPublic,
    QuizTestTakePublic,
    QuizTestCreate,
    QuizTestPublic,
    QuizAttemptPublic,
    QuizUserAnalyticsItem,
)

router = APIRouter(prefix="/tests", tags=["tests"])

IMPORT_HEADERS = [
    "Код_теста",
    "Название_теста",
    "Описание_теста",
    "ID_ресторана",
    "ID_роли",
    "Порядок_вопроса",
    "Текст_вопроса",
    "Тип_вопроса",
    "Количество_правильных_ответов",
    "Ответ_1",
    "Ответ_2",
    "Ответ_3",
    "Ответ_4",
    "Ответ_5",
    "Ответ_6",
    "Ответ_7",
    "Ответ_8",
]


def _normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return " ".join(value.strip().lower().replace("ё", "е").split())


def _validate_questions(payload: QuizTestCreate) -> None:
    for q_idx, question in enumerate(payload.questions, start=1):
        correct_count = sum(1 for opt in question.options if opt.is_correct)
        if question.question_type == QuestionType.SINGLE and correct_count != 1:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"В вопросе {q_idx} для single должен быть ровно один правильный вариант",
            )
        if question.question_type == QuestionType.MULTIPLE and correct_count < 1:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"В вопросе {q_idx} для multiple должен быть хотя бы один правильный вариант",
            )


def _parse_uuid_or_400(value: str, field_name: str) -> UUID:
    try:
        return UUID(value)
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Некорректный {field_name}") from exc


def _validate_restaurant_role(
    db: Session,
    restaurant_id: UUID,
    job_title_id: UUID,
) -> tuple[RestaurantCatalog, JobTitleCatalog]:
    restaurant = db.get(RestaurantCatalog, restaurant_id)
    if not restaurant:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Ресторан не найден")

    job_title = db.get(JobTitleCatalog, job_title_id)
    if not job_title:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Роль не найдена")
    if job_title.restaurant_id != restaurant.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Выбранная роль не принадлежит выбранному ресторану",
        )
    return restaurant, job_title


def _resolve_assignments(
    db: Session,
    payload: QuizTestCreate,
) -> list[tuple[RestaurantCatalog, JobTitleCatalog]]:
    """Собрать пары (ресторан, роль) из payload.assignments или из одиночных полей (legacy).

    Возвращает провалидированный список без дубликатов, сохраняя порядок.
    Первый элемент используется как «первичная» пара теста.
    """
    raw_pairs: list[tuple[str, str]] = []
    if payload.assignments:
        raw_pairs = [(a.restaurant_id, a.job_title_id) for a in payload.assignments]
    elif payload.restaurant_id and payload.job_title_id:
        raw_pairs = [(payload.restaurant_id, payload.job_title_id)]

    if not raw_pairs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Выберите хотя бы одну пару «ресторан + роль»",
        )

    resolved: list[tuple[RestaurantCatalog, JobTitleCatalog]] = []
    seen: set[tuple[UUID, UUID]] = set()
    for restaurant_id_raw, job_title_id_raw in raw_pairs:
        restaurant_id = _parse_uuid_or_400(restaurant_id_raw, "id ресторана")
        job_title_id = _parse_uuid_or_400(job_title_id_raw, "id роли")
        key = (restaurant_id, job_title_id)
        if key in seen:
            continue
        seen.add(key)
        resolved.append(_validate_restaurant_role(db, restaurant_id, job_title_id))
    return resolved


def _sync_test_assignments(
    db: Session,
    test: QuizTest,
    pairs: list[tuple[RestaurantCatalog, JobTitleCatalog]],
) -> None:
    """Привести строки quiz_test_assignments в соответствие со списком pairs."""
    wanted = {(restaurant.id, job_title.id) for restaurant, job_title in pairs}
    existing = list(db.scalars(select(QuizTestAssignment).where(QuizTestAssignment.test_id == test.id)).all())
    existing_keys = {(a.restaurant_id, a.job_title_id) for a in existing}

    for assignment in existing:
        if (assignment.restaurant_id, assignment.job_title_id) not in wanted:
            db.delete(assignment)
    for restaurant, job_title in pairs:
        if (restaurant.id, job_title.id) not in existing_keys:
            db.add(QuizTestAssignment(test_id=test.id, restaurant_id=restaurant.id, job_title_id=job_title.id))


def _upsert_test(
    db: Session,
    payload: QuizTestCreate,
    created_by_user_id: UUID,
    existing_test: QuizTest | None = None,
) -> tuple[QuizTest, str]:
    pairs = _resolve_assignments(db, payload)
    primary_restaurant, primary_job_title = pairs[0]
    _validate_questions(payload)

    external_code = payload.external_code.strip() if payload.external_code else None
    if external_code:
        duplicate = db.scalar(select(QuizTest).where(QuizTest.external_code == external_code))
        if duplicate and (existing_test is None or duplicate.id != existing_test.id):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Код теста уже существует")

    mode = "updated" if existing_test else "created"
    test = existing_test or QuizTest(created_by_user_id=created_by_user_id)
    test.external_code = external_code
    test.title = payload.title.strip()
    test.description = payload.description.strip() if payload.description else None
    test.restaurant_id = primary_restaurant.id
    test.job_title_id = primary_job_title.id

    if existing_test is None:
        db.add(test)
        db.flush([test])
    else:
        existing_questions = list(db.scalars(select(QuizQuestion).where(QuizQuestion.test_id == test.id)).all())
        for question in existing_questions:
            options = list(db.scalars(select(QuizOption).where(QuizOption.question_id == question.id)).all())
            for option in options:
                db.delete(option)
            db.delete(question)
        db.flush()

    for q_idx, q in enumerate(payload.questions):
        question = QuizQuestion(
            test_id=test.id,
            text=q.text.strip(),
            question_type=q.question_type,
            sort_order=q_idx,
        )
        db.add(question)
        db.flush([question])

        for o_idx, option in enumerate(q.options):
            db.add(
                QuizOption(
                    question_id=question.id,
                    text=option.text.strip(),
                    is_correct=option.is_correct,
                    sort_order=o_idx,
                )
            )

    _sync_test_assignments(db, test, pairs)
    return test, mode


@router.post("", response_model=QuizTestPublic, status_code=status.HTTP_201_CREATED)
def create_test(
    payload: QuizTestCreate,
    current_user: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    test, _ = _upsert_test(db, payload, created_by_user_id=current_user.id)
    db.commit()
    return _build_test_public(db, test.id)


@router.put("/{test_id}", response_model=QuizTestPublic)
def update_test(
    test_id: int,
    payload: QuizTestCreate,
    _: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    test = db.get(QuizTest, test_id)
    if not test:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Тест не найден")

    _upsert_test(db, payload, created_by_user_id=test.created_by_user_id, existing_test=test)

    db.commit()
    return _build_test_public(db, test.id)


@router.delete("/{test_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_test(
    test_id: int,
    _: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    test = db.get(QuizTest, test_id)
    if not test:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Тест не найден")

    attempts = list(db.scalars(select(QuizAttempt).where(QuizAttempt.test_id == test.id)).all())
    for attempt in attempts:
        attempt_answers = list(db.scalars(select(QuizAttemptAnswer).where(QuizAttemptAnswer.attempt_id == attempt.id)).all())
        for attempt_answer in attempt_answers:
            db.delete(attempt_answer)
        db.delete(attempt)

    questions = list(db.scalars(select(QuizQuestion).where(QuizQuestion.test_id == test.id)).all())
    for question in questions:
        options = list(db.scalars(select(QuizOption).where(QuizOption.question_id == question.id)).all())
        for option in options:
            db.delete(option)
        db.delete(question)

    assignments = list(db.scalars(select(QuizTestAssignment).where(QuizTestAssignment.test_id == test.id)).all())
    for assignment in assignments:
        db.delete(assignment)

    db.delete(test)
    db.commit()


@router.get("", response_model=list[QuizTestPublic])
def list_tests(
    restaurant_id: str | None = Query(default=None),
    job_title_id: str | None = Query(default=None),
    _: User = Depends(require_roles(Role.SUPERADMIN, Role.ADMIN)),
    db: Session = Depends(get_db),
):
    query = select(QuizTest).order_by(QuizTest.created_at.desc())
    try:
        if restaurant_id or job_title_id:
            # Фильтр по таблице привязок: тест попадает в выборку, если у него
            # есть пара (ресторан, роль), удовлетворяющая фильтру.
            assignment_filter = select(QuizTestAssignment.test_id)
            if restaurant_id:
                assignment_filter = assignment_filter.where(
                    QuizTestAssignment.restaurant_id == UUID(restaurant_id)
                )
            if job_title_id:
                assignment_filter = assignment_filter.where(
                    QuizTestAssignment.job_title_id == UUID(job_title_id)
                )
            query = query.where(QuizTest.id.in_(assignment_filter))
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Некорректный фильтр ресторана/роли") from exc

    tests = list(db.scalars(query).all())
    return [_build_test_public(db, test.id) for test in tests]


def _query_tests_for_user(db: Session, user: User) -> list[QuizTest]:
    query = select(QuizTest).order_by(QuizTest.created_at.desc())
    if user.role != Role.LEARNER:
        return list(db.scalars(query).all())

    user_restaurant = _normalize_text(user.restaurant)
    user_job_title = _normalize_text(user.job_title)
    if not user_restaurant or not user_job_title:
        return []

    tests = list(db.scalars(query).all())
    matched: list[QuizTest] = []
    for test in tests:
        # Тест доступен, если ХОТЯ БЫ ОДНА его пара (ресторан, роль) совпадает с профилем.
        assignments = list(
            db.scalars(select(QuizTestAssignment).where(QuizTestAssignment.test_id == test.id)).all()
        )
        for assignment in assignments:
            restaurant = db.get(RestaurantCatalog, assignment.restaurant_id)
            job_title = db.get(JobTitleCatalog, assignment.job_title_id)
            if not restaurant or not job_title:
                continue
            if (
                _normalize_text(restaurant.name) == user_restaurant
                and _normalize_text(job_title.name) == user_job_title
            ):
                matched.append(test)
                break
    return matched


@router.get("/my", response_model=list[QuizTestPublic])
def list_my_tests(
    current_user: User = Depends(require_roles(Role.SUPERADMIN, Role.ADMIN, Role.LEARNER)),
    db: Session = Depends(get_db),
):
    tests = _query_tests_for_user(db, current_user)
    return [_build_test_public(db, test.id) for test in tests]


@router.get("/{test_id}/take", response_model=QuizTestTakePublic)
def take_test(
    test_id: int,
    current_user: User = Depends(require_roles(Role.SUPERADMIN, Role.ADMIN, Role.LEARNER)),
    db: Session = Depends(get_db),
):
    test = db.get(QuizTest, test_id)
    if not test:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Тест не найден")

    if current_user.role == Role.LEARNER:
        available_ids = {item.id for item in _query_tests_for_user(db, current_user)}
        if test_id not in available_ids:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Тест недоступен для вашей роли")

    restaurant = db.get(RestaurantCatalog, test.restaurant_id)
    job_title = db.get(JobTitleCatalog, test.job_title_id)
    questions = list(
        db.scalars(select(QuizQuestion).where(QuizQuestion.test_id == test.id).order_by(QuizQuestion.sort_order.asc())).all()
    )
    question_public: list[QuizQuestionTakePublic] = []
    for question in questions:
        options = list(
            db.scalars(
                select(QuizOption).where(QuizOption.question_id == question.id).order_by(QuizOption.sort_order.asc())
            ).all()
        )
        question_public.append(
            QuizQuestionTakePublic(
                id=question.id,
                text=question.text,
                question_type=question.question_type,
                sort_order=question.sort_order,
                options=[QuizOptionTakePublic(id=o.id, text=o.text, sort_order=o.sort_order) for o in options],
            )
        )

    random.shuffle(question_public)

    return QuizTestTakePublic(
        id=test.id,
        title=test.title,
        description=test.description,
        restaurant_name=restaurant.name if restaurant else "",
        job_title_name=job_title.name if job_title else "",
        questions=question_public,
    )


def _build_attempt_public(db: Session, attempt: QuizAttempt) -> QuizAttemptPublic:
    user = db.get(User, attempt.user_id)
    test = db.get(QuizTest, attempt.test_id)
    return QuizAttemptPublic(
        id=attempt.id,
        test_id=attempt.test_id,
        test_title=test.title if test else "",
        user_id=str(attempt.user_id),
        user_name=user.full_name if user else "",
        user_email=user.email if user else "",
        user_restaurant=user.restaurant if user else None,
        user_job_title=user.job_title if user else None,
        started_at=attempt.started_at,
        finished_at=attempt.finished_at,
        duration_seconds=attempt.duration_seconds,
        total_questions=attempt.total_questions,
        correct_answers=attempt.correct_answers,
        incorrect_answers=attempt.incorrect_answers,
    )


def _split_saved_options(value: str) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(" | ") if item.strip()]


@router.get("/my-attempts", response_model=list[QuizAttemptPublic])
def list_my_attempts(
    current_user: User = Depends(require_roles(Role.SUPERADMIN, Role.ADMIN, Role.LEARNER)),
    db: Session = Depends(get_db),
):
    attempts = list(
        db.scalars(
            select(QuizAttempt).where(QuizAttempt.user_id == current_user.id).order_by(QuizAttempt.finished_at.desc())
        ).all()
    )
    return [_build_attempt_public(db, attempt) for attempt in attempts]


@router.get("/my-attempts/{attempt_id}", response_model=QuizAttemptDetailPublic)
def get_my_attempt_detail(
    attempt_id: int,
    current_user: User = Depends(require_roles(Role.SUPERADMIN, Role.ADMIN, Role.LEARNER)),
    db: Session = Depends(get_db),
):
    attempt = db.get(QuizAttempt, attempt_id)
    if not attempt:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Попытка не найдена")
    if attempt.user_id != current_user.id and current_user.role == Role.LEARNER:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Нет доступа к этой попытке")
    answers = list(
        db.scalars(select(QuizAttemptAnswer).where(QuizAttemptAnswer.attempt_id == attempt_id).order_by(QuizAttemptAnswer.id.asc())).all()
    )
    detail_results = [
        QuizAttemptQuestionDetailPublic(
            question_id=answer.question_id,
            question_text=answer.question_text,
            selected_options=_split_saved_options(answer.selected_options_text),
            correct_options=_split_saved_options(answer.correct_options_text),
            is_correct=answer.is_correct,
        )
        for answer in answers
    ]
    rng = random.Random(attempt.id)
    rng.shuffle(detail_results)
    return QuizAttemptDetailPublic(
        attempt=_build_attempt_public(db, attempt),
        results=detail_results,
    )


@router.get("/scoreboard", response_model=QuizScoreboardResponse)
def tests_scoreboard(
    _: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    """Сводная таблица баллов: по каждому пользователю — лучший и последний
    результат по каждому тесту, который он проходил."""
    attempts = list(db.scalars(select(QuizAttempt).order_by(QuizAttempt.finished_at.asc())).all())

    def percent(attempt: QuizAttempt) -> float:
        if attempt.total_questions <= 0:
            return 0.0
        return attempt.correct_answers / attempt.total_questions * 100

    # (user_id, test_id) -> {"best": QuizAttempt, "last": QuizAttempt, "count": int}
    buckets: dict[tuple[UUID, int], dict] = {}
    for attempt in attempts:
        key = (attempt.user_id, attempt.test_id)
        bucket = buckets.setdefault(key, {"best": attempt, "last": attempt, "count": 0})
        bucket["count"] += 1
        bucket["last"] = attempt  # список отсортирован по finished_at asc
        if percent(attempt) > percent(bucket["best"]):
            bucket["best"] = attempt

    test_ids = {test_id for _, test_id in buckets}
    tests_by_id = {test_id: db.get(QuizTest, test_id) for test_id in test_ids}
    test_refs = sorted(
        (
            QuizScoreboardTestRef(id=test_id, title=test.title)
            for test_id, test in tests_by_id.items()
            if test
        ),
        key=lambda item: item.title.lower(),
    )

    users_cells: dict[UUID, list[QuizScoreboardCell]] = {}
    for (user_id, test_id), bucket in buckets.items():
        if not tests_by_id.get(test_id):
            continue
        best: QuizAttempt = bucket["best"]
        last: QuizAttempt = bucket["last"]
        users_cells.setdefault(user_id, []).append(
            QuizScoreboardCell(
                test_id=test_id,
                attempts_count=bucket["count"],
                best_correct=best.correct_answers,
                best_total=best.total_questions,
                best_percent=percent(best),
                last_correct=last.correct_answers,
                last_total=last.total_questions,
                last_percent=percent(last),
                last_finished_at=last.finished_at,
            )
        )

    scoreboard_users: list[QuizScoreboardUser] = []
    for user_id, cells in users_cells.items():
        user = db.get(User, user_id)
        scoreboard_users.append(
            QuizScoreboardUser(
                user_id=str(user_id),
                user_name=user.full_name if user else "",
                user_email=user.email if user else "",
                user_restaurant=user.restaurant if user else None,
                user_job_title=user.job_title if user else None,
                scores=cells,
            )
        )
    scoreboard_users.sort(key=lambda item: item.user_name.lower())

    return QuizScoreboardResponse(tests=test_refs, users=scoreboard_users)


@router.get("/analytics", response_model=QuizAnalyticsResponse)
def tests_analytics(
    limit_recent: int = Query(default=5, ge=1, le=50),
    attempts_limit: int = Query(default=200, ge=1, le=1000),
    attempts_offset: int = Query(default=0, ge=0),
    _: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    all_attempts = list(db.scalars(select(QuizAttempt).order_by(QuizAttempt.finished_at.desc())).all())
    recent_attempts = all_attempts[:limit_recent]
    attempts_page = all_attempts[attempts_offset : attempts_offset + attempts_limit]

    unique_users = len({attempt.user_id for attempt in all_attempts})
    avg_score = 0.0
    if all_attempts:
        avg_score = sum(
            (attempt.correct_answers / attempt.total_questions) * 100
            for attempt in all_attempts
            if attempt.total_questions > 0
        ) / len(all_attempts)

    durations = [attempt.duration_seconds for attempt in all_attempts if attempt.duration_seconds is not None]
    avg_duration = (sum(durations) / len(durations)) if durations else None

    answers = list(db.scalars(select(QuizAttemptAnswer)).all())

    question_buckets: dict[int, dict] = {}
    for answer in answers:
        if answer.question_id is None:
            continue
        bucket = question_buckets.setdefault(
            answer.question_id,
            {
                "question_text": answer.question_text,
                "total_attempts": 0,
                "wrong_attempts": 0,
            },
        )
        bucket["total_attempts"] += 1
        if not answer.is_correct:
            bucket["wrong_attempts"] += 1

    question_analytics: list[QuizQuestionAnalyticsItem] = []
    for question_id, bucket in question_buckets.items():
        question = db.get(QuizQuestion, question_id)
        if not question:
            continue
        test = db.get(QuizTest, question.test_id)
        total_attempts = bucket["total_attempts"]
        wrong_attempts = bucket["wrong_attempts"]
        correct_attempts = total_attempts - wrong_attempts
        wrong_rate = (wrong_attempts / total_attempts) if total_attempts else 0.0
        question_analytics.append(
            QuizQuestionAnalyticsItem(
                question_id=question_id,
                test_id=question.test_id,
                test_title=test.title if test else "",
                question_text=bucket["question_text"],
                total_attempts=total_attempts,
                wrong_attempts=wrong_attempts,
                correct_attempts=correct_attempts,
                wrong_rate=wrong_rate,
            )
        )
    question_analytics.sort(key=lambda item: (-item.wrong_rate, -item.wrong_attempts, item.question_id))

    user_answers_bucket: dict[UUID, dict] = {}
    for answer in answers:
        attempt = db.get(QuizAttempt, answer.attempt_id)
        if not attempt:
            continue
        bucket = user_answers_bucket.setdefault(
            attempt.user_id,
            {"total_answers": 0, "wrong_answers": 0, "durations": [], "attempt_ids": set()},
        )
        bucket["total_answers"] += 1
        bucket["wrong_answers"] += 0 if answer.is_correct else 1
        bucket["attempt_ids"].add(attempt.id)
        if attempt.duration_seconds is not None:
            bucket["durations"].append(attempt.duration_seconds)

    user_analytics: list[QuizUserAnalyticsItem] = []
    for user_id, bucket in user_answers_bucket.items():
        user = db.get(User, user_id)
        total_answers = bucket["total_answers"]
        wrong_answers = bucket["wrong_answers"]
        wrong_rate = (wrong_answers / total_answers) if total_answers else 0.0
        durations_for_user = bucket["durations"]
        avg_user_duration = (sum(durations_for_user) / len(durations_for_user)) if durations_for_user else None
        user_analytics.append(
            QuizUserAnalyticsItem(
                user_id=str(user_id),
                user_name=user.full_name if user else "",
                user_email=user.email if user else "",
                attempts_count=len(bucket["attempt_ids"]),
                total_answers=total_answers,
                wrong_answers=wrong_answers,
                wrong_rate=wrong_rate,
                avg_duration_seconds=avg_user_duration,
            )
        )
    user_analytics.sort(key=lambda item: (-item.wrong_answers, -item.wrong_rate, item.user_name))

    return QuizAnalyticsResponse(
        summary=QuizAnalyticsSummary(
            total_attempts=len(all_attempts),
            unique_users=unique_users,
            avg_score_percent=avg_score,
            avg_duration_seconds=avg_duration,
        ),
        recent_attempts=[_build_attempt_public(db, attempt) for attempt in recent_attempts],
        attempts=[_build_attempt_public(db, attempt) for attempt in attempts_page],
        question_analytics=question_analytics,
        user_analytics=user_analytics,
    )


@router.post("/{test_id}/submit", response_model=QuizSubmitResultPublic)
def submit_test(
    test_id: int,
    payload: QuizSubmitRequest,
    current_user: User = Depends(require_roles(Role.SUPERADMIN, Role.ADMIN, Role.LEARNER)),
    db: Session = Depends(get_db),
):
    _ = take_test(test_id=test_id, current_user=current_user, db=db)
    questions = list(
        db.scalars(select(QuizQuestion).where(QuizQuestion.test_id == test_id).order_by(QuizQuestion.sort_order.asc())).all()
    )
    answers_by_question = {item.question_id: set(item.option_ids) for item in payload.answers}

    results: list[QuizQuestionResultPublic] = []
    correct_answers = 0
    finished_at = datetime.utcnow()
    started_at = payload.started_at
    if started_at and started_at.tzinfo is not None:
        started_at = started_at.astimezone(timezone.utc).replace(tzinfo=None)
    duration_seconds: int | None = None
    if started_at and started_at <= finished_at:
        duration_seconds = int((finished_at - started_at).total_seconds())
    attempt_answers_payload: list[dict] = []

    for question in questions:
        options = list(
            db.scalars(
                select(QuizOption).where(QuizOption.question_id == question.id).order_by(QuizOption.sort_order.asc())
            ).all()
        )
        option_text_by_id = {option.id: option.text for option in options}
        correct_ids = {option.id for option in options if option.is_correct}
        selected_ids = answers_by_question.get(question.id, set())
        ordered_correct_ids = [option.id for option in options if option.id in correct_ids]
        ordered_selected_ids = [option.id for option in options if option.id in selected_ids]

        is_correct = selected_ids == correct_ids
        if is_correct:
            correct_answers += 1
        correct_texts = [option_text_by_id[option_id] for option_id in ordered_correct_ids if option_id in option_text_by_id]
        selected_texts = [option_text_by_id[option_id] for option_id in ordered_selected_ids if option_id in option_text_by_id]

        attempt_answers_payload.append(
            {
                "question_id": question.id,
                "question_text": question.text,
                "selected_option_ids": ",".join(str(option_id) for option_id in ordered_selected_ids),
                "selected_options_text": " | ".join(selected_texts),
                "correct_option_ids": ",".join(str(option_id) for option_id in ordered_correct_ids),
                "correct_options_text": " | ".join(correct_texts),
                "is_correct": is_correct,
            }
        )

        results.append(
            QuizQuestionResultPublic(
                question_id=question.id,
                question_text=question.text,
                correct_options=correct_texts,
                selected_options=selected_texts,
                is_correct=is_correct,
            )
        )

    random.shuffle(results)

    total_questions = len(questions)
    attempt = QuizAttempt(
        test_id=test_id,
        user_id=current_user.id,
        started_at=started_at,
        finished_at=finished_at,
        duration_seconds=duration_seconds,
        total_questions=total_questions,
        correct_answers=correct_answers,
        incorrect_answers=total_questions - correct_answers,
    )
    db.add(attempt)
    db.flush([attempt])
    for answer_payload in attempt_answers_payload:
        db.add(QuizAttemptAnswer(attempt_id=attempt.id, **answer_payload))
    db.commit()

    return QuizSubmitResultPublic(
        attempt_id=attempt.id,
        started_at=started_at,
        finished_at=finished_at,
        duration_seconds=duration_seconds,
        total_questions=total_questions,
        correct_answers=correct_answers,
        incorrect_answers=total_questions - correct_answers,
        results=results,
    )


@router.get("/import-template")
def download_import_template(
    _: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт"
    ws.append(IMPORT_HEADERS)
    restaurants = list(db.scalars(select(RestaurantCatalog).order_by(RestaurantCatalog.name.asc())).all())
    first_restaurant_id = ""
    first_role_id = ""
    if restaurants:
        first_restaurant_id = str(restaurants[0].id)
        first_role = db.scalar(
            select(JobTitleCatalog)
            .where(JobTitleCatalog.restaurant_id == restaurants[0].id)
            .order_by(JobTitleCatalog.name.asc())
            .limit(1)
        )
        if first_role:
            first_role_id = str(first_role.id)

    ws.append(
        [
            "waiter_basic_01",
            "Базовый тест официанта",
            "Проверка базовых знаний меню",
            first_restaurant_id,
            first_role_id,
            1,
            "Какой напиток подается холодным?",
            "single",
            1,
            "Лимонад",
            "Латте",
            "Капучино",
            "",
            "",
            "",
            "",
            "",
        ]
    )

    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, len(IMPORT_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    column_widths = {
        "A": 18, "B": 28, "C": 36, "D": 28, "E": 28,
        "F": 16, "G": 42, "H": 14, "I": 24,
        "J": 20, "K": 20, "L": 20, "M": 20,
        "N": 20, "O": 20, "P": 20, "Q": 20,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    ref_ws = wb.create_sheet("Справочники")
    ref_ws.append(["ID_ресторана", "Ресторан", "ID_роли", "Роль"])
    for col_idx in range(1, 5):
        cell = ref_ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ref_ws.column_dimensions["A"].width = 40
    ref_ws.column_dimensions["B"].width = 28
    ref_ws.column_dimensions["C"].width = 40
    ref_ws.column_dimensions["D"].width = 24
    ref_ws.freeze_panes = "A2"
    for restaurant in restaurants:
        roles = list(
            db.scalars(
                select(JobTitleCatalog)
                .where(JobTitleCatalog.restaurant_id == restaurant.id)
                .order_by(JobTitleCatalog.name.asc())
            ).all()
        )
        if not roles:
            ref_ws.append([str(restaurant.id), restaurant.name, "", ""])
            continue
        for role in roles:
            ref_ws.append([str(restaurant.id), restaurant.name, str(role.id), role.name])

    rules = wb.create_sheet("Правила")
    rules.append(["Правила заполнения"])
    rules["A1"].font = Font(bold=True)
    rules.append(["1) Одна строка = один вопрос."])
    rules.append(["2) Тип_вопроса: single или multiple."])
    rules.append(["3) Количество_правильных_ответов = N, первые N ответов считаются правильными."])
    rules.append(["4) Остальные ответы автоматически считаются неправильными."])
    rules.append(["5) Для single должно быть N=1, для multiple N>=1."])
    rules.append(["6) Код_теста объединяет строки вопросов в один тест."])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tests_import_template.xlsx"},
    )


@router.post("/import-xlsx")
async def import_tests_from_xlsx(
    file: UploadFile = File(...),
    dry_run: bool = Query(default=True),
    current_user: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Файл пустой")

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Не удалось прочитать Excel: {exc}") from exc

    ws = wb["Импорт"] if "Импорт" in wb.sheetnames else wb.active
    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        header_map[str(cell.value).strip()] = col_idx

    missing_headers = [h for h in IMPORT_HEADERS if h not in header_map]
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"В файле отсутствуют колонки: {', '.join(missing_headers)}",
        )

    grouped: dict[str, dict] = {}
    errors: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        def get_val(header: str):
            return ws.cell(row=row_idx, column=header_map[header]).value

        code = str(get_val("Код_теста") or "").strip()
        title = str(get_val("Название_теста") or "").strip()
        description = str(get_val("Описание_теста") or "").strip() or None
        restaurant_id_raw = str(get_val("ID_ресторана") or "").strip()
        role_id_raw = str(get_val("ID_роли") or "").strip()
        question_text = str(get_val("Текст_вопроса") or "").strip()
        question_type_raw = str(get_val("Тип_вопроса") or "").strip().lower()
        correct_count_raw = get_val("Количество_правильных_ответов")

        if not any([code, title, question_text]):
            continue

        if not code or not title or not question_text:
            errors.append(f"Строка {row_idx}: заполните Код_теста, Название_теста и Текст_вопроса")
            continue

        try:
            UUID(restaurant_id_raw)
            UUID(role_id_raw)
        except ValueError:
            errors.append(f"Строка {row_idx}: некорректный ID_ресторана или ID_роли")
            continue

        if question_type_raw not in {QuestionType.SINGLE.value, QuestionType.MULTIPLE.value}:
            errors.append(f"Строка {row_idx}: Тип_вопроса должен быть single или multiple")
            continue

        try:
            correct_count = int(correct_count_raw)
        except (TypeError, ValueError):
            errors.append(f"Строка {row_idx}: Количество_правильных_ответов должно быть числом")
            continue

        answers: list[str] = []
        for i in range(1, 9):
            ans = str(get_val(f"Ответ_{i}") or "").strip()
            if ans:
                answers.append(ans)

        if len(answers) < 2:
            errors.append(f"Строка {row_idx}: должно быть минимум 2 непустых варианта ответа")
            continue
        if correct_count < 1 or correct_count > len(answers):
            errors.append(f"Строка {row_idx}: Количество_правильных_ответов должно быть от 1 до {len(answers)}")
            continue
        if question_type_raw == QuestionType.SINGLE.value and correct_count != 1:
            errors.append(f"Строка {row_idx}: для single количество правильных должно быть 1")
            continue

        options = [{"text": text, "is_correct": idx < correct_count} for idx, text in enumerate(answers)]

        if code not in grouped:
            grouped[code] = {
                "external_code": code,
                "title": title,
                "description": description,
                "restaurant_id": restaurant_id_raw,
                "job_title_id": role_id_raw,
                "questions": [],
            }
        else:
            group = grouped[code]
            if (
                group["title"] != title
                or group["restaurant_id"] != restaurant_id_raw
                or group["job_title_id"] != role_id_raw
            ):
                errors.append(
                    f"Строка {row_idx}: у теста {code} должны совпадать название, ресторан и роль во всех строках"
                )
                continue

        grouped[code]["questions"].append(
            {
                "text": question_text,
                "question_type": question_type_raw,
                "options": options,
            }
        )

    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=errors)

    created = 0
    updated = 0
    for code, payload_data in grouped.items():
        payload = QuizTestCreate.model_validate(payload_data)
        existing = db.scalar(select(QuizTest).where(QuizTest.external_code == code))
        _, mode = _upsert_test(
            db,
            payload=payload,
            created_by_user_id=current_user.id,
            existing_test=existing,
        )
        if mode == "created":
            created += 1
        else:
            updated += 1

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return {
        "dry_run": dry_run,
        "tests_in_file": len(grouped),
        "created": created,
        "updated": updated,
        "errors": [],
    }


@router.post("/import-docx")
async def import_tests_from_docx(
    file: UploadFile = File(...),
    restaurant_id: str = Form(...),
    job_title_id: str = Form(...),
    title: str = Form(...),
    external_code: str = Form(...),
    description: str | None = Form(default=None),
    dry_run: bool = Query(default=True),
    current_user: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    """
    Импорт одного теста из .docx: вопросы в нумерованном виде «1.», варианты «A)», «B)», …;
    правильные ответы выделены жирным в Word.
    """
    name = (file.filename or "").lower()
    if not name.endswith(".docx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Нужен файл .docx (Word)",
        )
    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Файл пустой")

    try:
        UUID(restaurant_id)
        UUID(job_title_id)
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Некорректный ID ресторана или роли",
        ) from exc

    questions, errors = parse_docx_to_questions(content)
    if errors:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=errors)

    payload = QuizTestCreate.model_validate(
        {
            "external_code": external_code.strip(),
            "title": title.strip(),
            "description": description.strip() if description else None,
            "restaurant_id": restaurant_id,
            "job_title_id": job_title_id,
            "questions": questions,
        }
    )

    existing = db.scalar(select(QuizTest).where(QuizTest.external_code == payload.external_code))
    _, mode = _upsert_test(
        db,
        payload=payload,
        created_by_user_id=current_user.id,
        existing_test=existing,
    )

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return {
        "dry_run": dry_run,
        "tests_in_file": 1,
        "created": 1 if mode == "created" else 0,
        "updated": 1 if mode == "updated" else 0,
        "mode": mode,
        "questions_count": len(questions),
        "errors": [],
    }


@router.post("/parse-docx")
async def parse_docx_preview(
    file: UploadFile = File(...),
    _: User = Depends(require_roles(Role.SUPERADMIN)),
):
    """Только разбор .docx без записи в БД — для предпросмотра и правок на фронте.

    Лояльный режим: проблемные вопросы не блокируют разбор, а возвращаются
    вместе с предупреждениями — их правят в предпросмотре перед сохранением.
    """
    name = (file.filename or "").lower()
    if not name.endswith(".docx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Нужен файл .docx (Word)",
        )
    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Файл пустой")

    questions, warnings = parse_docx_to_questions(content, strict=False)
    if not questions:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=warnings)

    return {"questions": questions, "warnings": warnings}

@router.post("/import-apply")
def import_apply_test_payload(
    payload: QuizTestCreate,
    dry_run: bool = Query(default=False),
    current_user: User = Depends(require_roles(Role.SUPERADMIN)),
    db: Session = Depends(get_db),
):
    """
    Сохранение теста после предпросмотра/редактирования (тот же upsert по external_code, что и import-docx).
    """
    if not payload.external_code or not str(payload.external_code).strip():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Укажите код теста (external_code)",
        )

    existing = db.scalar(
        select(QuizTest).where(QuizTest.external_code == str(payload.external_code).strip())
    )
    _, mode = _upsert_test(
        db,
        payload=payload,
        created_by_user_id=current_user.id,
        existing_test=existing,
    )

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return {
        "dry_run": dry_run,
        "created": 1 if mode == "created" else 0,
        "updated": 1 if mode == "updated" else 0,
        "mode": mode,
        "questions_count": len(payload.questions),
    }


def _build_test_public(db: Session, test_id: int) -> QuizTestPublic:
    test = db.get(QuizTest, test_id)
    if not test:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Тест не найден")

    restaurant = db.get(RestaurantCatalog, test.restaurant_id)
    job_title = db.get(JobTitleCatalog, test.job_title_id)

    assignment_rows = list(
        db.scalars(select(QuizTestAssignment).where(QuizTestAssignment.test_id == test.id)).all()
    )
    assignments: list[QuizTestAssignmentPublic] = []
    for row in assignment_rows:
        a_restaurant = db.get(RestaurantCatalog, row.restaurant_id)
        a_job_title = db.get(JobTitleCatalog, row.job_title_id)
        if not a_restaurant or not a_job_title:
            continue
        assignments.append(
            QuizTestAssignmentPublic(
                restaurant_id=str(row.restaurant_id),
                restaurant_name=a_restaurant.name,
                job_title_id=str(row.job_title_id),
                job_title_name=a_job_title.name,
            )
        )

    questions = list(db.scalars(select(QuizQuestion).where(QuizQuestion.test_id == test.id).order_by(QuizQuestion.sort_order.asc())).all())
    question_public: list[QuizQuestionPublic] = []
    for question in questions:
        options = list(
            db.scalars(
                select(QuizOption)
                .where(QuizOption.question_id == question.id)
                .order_by(QuizOption.sort_order.asc())
            ).all()
        )
        question_public.append(
            QuizQuestionPublic(
                id=question.id,
                text=question.text,
                question_type=question.question_type,
                sort_order=question.sort_order,
                options=[QuizOptionPublic.model_validate(o) for o in options],
            )
        )

    return QuizTestPublic(
        id=test.id,
        external_code=test.external_code,
        title=test.title,
        description=test.description,
        restaurant_id=str(test.restaurant_id),
        restaurant_name=restaurant.name if restaurant else "",
        job_title_id=str(test.job_title_id),
        job_title_name=job_title.name if job_title else "",
        assignments=assignments,
        created_at=test.created_at,
        questions=question_public,
    )
