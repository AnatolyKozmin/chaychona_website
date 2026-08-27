"""Разбор файла-реестра «Вкусной тетради», который заливают через админку.

Принимаем либо голый .xlsx (только текст), либо .zip, внутри которого лежит
.xlsx и папки с медиа (`dish_photos/`, `images/`, `audio/`) — пути к файлам
берутся из самих колонок реестра. Zip нужен потому, что через браузер нельзя
отдать «относительные пути с диска»: файлы должны приехать вместе с реестром.

Формат листа — тот же, что у ручного скрипта `scripts/import_from_registry.py`:

    №  |  Раздел  |  Блюдо  |  Ингредиенты (подписи)  |  Текст озвучки  |
    Фото блюда (файл)  |  Картинка ингредиентов (файл)  |  Озвучка (файл)

Колонки ищем по заголовку, а не по номеру: заказчик двигает столбцы и
дописывает свои. Если заголовок не опознан — откатываемся на порядок выше.
"""
from __future__ import annotations

import io
import zipfile
from dataclasses import dataclass, field
from pathlib import Path, PurePosixPath

import openpyxl

# Позиции по умолчанию — если заголовки не опознались совсем.
_FALLBACK_ORDER = [
    "number",
    "category",
    "name",
    "ingredients",
    "description",
    "photo_dish",
    "photo_ingredients",
    "audio",
]

# Пустая ячейка: реестр допускает тире вместо пропуска.
_EMPTY_VALUES = {"", "—", "–", "-", "none", "nan"}


class RegistryParseError(RuntimeError):
    """Файл не читается как реестр: не тот формат, нет листа, нет колонки блюда."""


@dataclass
class ParsedRow:
    """Одна строка реестра после разбора."""

    row_number: int
    name: str
    category: str | None = None
    ingredients: str | None = None
    description: str | None = None
    price_rubles: str | None = None
    photo_dish: str | None = None
    photo_ingredients: str | None = None
    audio: str | None = None


@dataclass
class ParsedRegistry:
    """Разобранный реестр вместе с доступом к медиа из архива."""

    rows: list[ParsedRow]
    source_name: str
    archive: zipfile.ZipFile | None = None
    media_index: dict[str, str] = field(default_factory=dict)

    def read_media(self, relpath: str | None) -> bytes | None:
        """Достать файл из архива по пути из реестра.

        Ищем сначала по полному пути, потом по одному имени файла: заказчик
        нередко пакует папки на уровень глубже, чем прописано в Excel.
        """
        if not relpath or self.archive is None:
            return None
        for key in _media_keys(relpath):
            member = self.media_index.get(key)
            if member is not None:
                with self.archive.open(member) as handle:
                    return handle.read()
        return None

    def close(self) -> None:
        if self.archive is not None:
            self.archive.close()
            self.archive = None


def _norm(value: object) -> str | None:
    """Ячейка → строка или None. Тире и пробелы считаем пустотой."""
    if value is None:
        return None
    text = " ".join(str(value).split())
    if text.lower() in _EMPTY_VALUES:
        return None
    return text


def _media_keys(path: str) -> list[str]:
    """Ключи, по которым ищем файл: полный путь и голое имя, всё в нижнем регистре."""
    normalized = path.strip().replace("\\", "/").lstrip("./").lower()
    if not normalized:
        return []
    keys = [normalized]
    name = PurePosixPath(normalized).name
    if name and name != normalized:
        keys.append(name)
    return keys


def _member_names(info: zipfile.ZipInfo) -> list[str]:
    """Имя члена архива во всех кодировках, в которых оно может быть записано.

    Бит 0x800 означает UTF-8. Без него zipfile декодирует имя как CP437, и
    русские названия превращаются в мусор — исходные байты надо перечитать
    как CP866, иначе пути из Excel никогда не совпадут.
    """
    names = [info.filename]
    if not info.flag_bits & 0x800:
        try:
            names.append(info.filename.encode("cp437").decode("cp866"))
        except (UnicodeEncodeError, UnicodeDecodeError):
            pass
    return names


def _detect_columns(header: tuple) -> dict[str, int]:
    """Сопоставить заголовки реестра с нашими полями."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(header):
        title = (_norm(raw) or "").lower()
        if not title:
            continue
        is_file = "файл" in title
        if "блюдо" in title and not is_file:
            mapping.setdefault("name", idx)
        elif "раздел" in title or "категор" in title:
            mapping.setdefault("category", idx)
        elif "ингредиент" in title and not is_file and "картинк" not in title:
            mapping.setdefault("ingredients", idx)
        elif "озвучк" in title and "текст" in title:
            mapping.setdefault("description", idx)
        elif "описан" in title:
            mapping.setdefault("description", idx)
        elif "цена" in title or "стоимост" in title:
            mapping.setdefault("price_rubles", idx)
        elif "фото блюда" in title:
            mapping.setdefault("photo_dish", idx)
        elif "картинк" in title and "ингредиент" in title and "magnific" not in title:
            mapping.setdefault("photo_ingredients", idx)
        elif "озвучк" in title and "magnific" not in title:
            mapping.setdefault("audio", idx)
    return mapping


def _rows_from_sheet(sheet) -> list[ParsedRow]:
    raw_rows = list(sheet.iter_rows(values_only=True))
    if not raw_rows:
        raise RegistryParseError("В файле нет ни одной строки")

    columns = _detect_columns(raw_rows[0])
    if "name" not in columns:
        # Заголовок не опознан — считаем порядок колонок каноническим.
        columns = {key: idx for idx, key in enumerate(_FALLBACK_ORDER)}

    rows: list[ParsedRow] = []
    for offset, raw in enumerate(raw_rows[1:], start=2):

        def cell(key: str, row=raw) -> str | None:
            idx = columns.get(key)
            if idx is None or idx >= len(row):
                return None
            return _norm(row[idx])

        name = cell("name")
        if not name:
            continue  # пустая строка-разделитель, а не ошибка
        rows.append(
            ParsedRow(
                row_number=offset,
                name=name,
                category=cell("category"),
                ingredients=cell("ingredients"),
                description=cell("description"),
                price_rubles=cell("price_rubles"),
                photo_dish=cell("photo_dish"),
                photo_ingredients=cell("photo_ingredients"),
                audio=cell("audio"),
            )
        )
    if not rows:
        raise RegistryParseError("В реестре не нашлось ни одного блюда — проверьте колонку «Блюдо»")
    return rows


def _load_workbook_rows(data: bytes) -> list[ParsedRow]:
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — openpyxl кидает разное на битых файлах
        raise RegistryParseError(f"Не удалось прочитать Excel: {exc}") from exc
    try:
        return _rows_from_sheet(workbook.active)
    finally:
        workbook.close()


def parse_registry(source: bytes | str | Path, file_name: str) -> ParsedRegistry:
    """Разобрать загруженный файл: .xlsx напрямую или .zip с реестром и медиа.

    `source` — либо байты, либо путь к файлу на диске. Путь нужен для больших
    архивов: залив на сотню блюд с фотографиями весит сотни мегабайт, и держать
    его целиком в памяти воркера незачем — zipfile прекрасно читает с диска.
    """
    lowered = (file_name or "").lower()
    handle: bytes | str = source if isinstance(source, bytes) else str(source)

    if lowered.endswith(".zip"):
        try:
            archive = zipfile.ZipFile(io.BytesIO(handle) if isinstance(handle, bytes) else handle)
        except zipfile.BadZipFile as exc:
            raise RegistryParseError("Архив повреждён или это не zip") from exc

        media_index: dict[str, str] = {}
        workbook_member: str | None = None
        for info in archive.infolist():
            if info.is_dir():
                continue
            for name in _member_names(info):
                base = PurePosixPath(name).name
                # ~$-файлы — временные копии Word/Excel, не реестр.
                if name.lower().endswith((".xlsx", ".xlsm")) and not base.startswith("~$"):
                    workbook_member = workbook_member or info.filename
                for key in _media_keys(name):
                    media_index.setdefault(key, info.filename)

        if workbook_member is None:
            archive.close()
            raise RegistryParseError("В архиве нет файла .xlsx с реестром")

        try:
            rows = _load_workbook_rows(archive.read(workbook_member))
        except Exception:
            archive.close()
            raise
        return ParsedRegistry(
            rows=rows,
            source_name=file_name,
            archive=archive,
            media_index=media_index,
        )

    if lowered.endswith((".xlsx", ".xlsm")):
        data = handle if isinstance(handle, bytes) else Path(handle).read_bytes()
        return ParsedRegistry(rows=_load_workbook_rows(data), source_name=file_name)

    raise RegistryParseError("Поддерживаются только .xlsx и .zip")
